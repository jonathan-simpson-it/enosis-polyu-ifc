"""Generate test trade documents for demo and testing."""
import csv
import io
import os
from pathlib import Path

DIR = Path(__file__).parent / "test-docs"
DIR.mkdir(parents=True, exist_ok=True)


def generate_invoice_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        print("reportlab not installed, skipping PDF generation")
        return None

    path = DIR / "invoice-sample.pdf"
    doc = SimpleDocTemplate(str(path), pagesize=A4)
    styles = getSampleStyleSheet()

    elements = []
    elements.append(Paragraph("INVOICE", styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("INVOICE #: INV-2026-0715-0042", styles["Normal"]))
    elements.append(Paragraph("DATE: 15-JUL-2026", styles["Normal"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("<b>CONSIGNOR:</b>", styles["Normal"]))
    elements.append(Paragraph("Shenzhen Electronics Trading Co., Ltd.", styles["Normal"]))
    elements.append(Paragraph("15 Huaqiang Bei Road, Futian District", styles["Normal"]))
    elements.append(Paragraph("Shenzhen, Guangdong, 518000", styles["Normal"]))
    elements.append(Spacer(1, 8))

    elements.append(Paragraph("<b>CONSIGNEE:</b>", styles["Normal"]))
    elements.append(Paragraph("HK Digital Logistics Ltd.", styles["Normal"]))
    elements.append(Paragraph("22/F Enterprise Building", styles["Normal"]))
    elements.append(Paragraph("118 Connaught Road West, Hong Kong", styles["Normal"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("<b>SHIPMENT DETAILS</b>", styles["Heading2"]))
    elements.append(Paragraph("Container: MSCU4820137", styles["Normal"]))
    elements.append(Paragraph("Port of Loading: Yantian, Shenzhen", styles["Normal"]))
    elements.append(Paragraph("Port of Discharge: Hong Kong", styles["Normal"]))
    elements.append(Paragraph("Vessel: MSC GENEVA V.126W", styles["Normal"]))
    elements.append(Paragraph("Incoterms: CIF Hong Kong", styles["Normal"]))
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("<b>COMMODITIES</b>", styles["Heading2"]))
    data = [
        ["#", "Description", "HS Code", "Qty", "Unit Price", "Total Value", "Weight (kg)"],
        ["1", "Portable laptop computers, Lenovo ThinkPad X1", "8471.30.00", "500", "USD 850.00", "USD 425,000.00", "750.0"],
        ["2", "Solid-state storage drives, 1TB NVMe", "8523.51.00", "1000", "USD 89.00", "USD 89,000.00", "120.0"],
        ["3", "Electronic integrated circuits, processor controllers", "8542.31.00", "10000", "USD 12.50", "USD 125,000.00", "45.0"],
    ]
    table = Table(data, colWidths=[24, 180, 70, 50, 60, 70, 60])
    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.9, 0.9, 0.9)),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ])
    table.setStyle(style)
    elements.append(table)
    elements.append(Spacer(1, 12))

    elements.append(Paragraph("<b>TOTAL DECLARED VALUE: USD 639,000.00</b>", styles["Normal"]))
    elements.append(Paragraph("TOTAL GROSS WEIGHT: 915.0 kg", styles["Normal"]))
    elements.append(Paragraph("TOTAL NET WEIGHT: 832.0 kg", styles["Normal"]))
    elements.append(Paragraph("NUMBER OF PACKAGES: 28 pallets", styles["Normal"]))

    doc.build(elements)
    print(f"Generated: {path}")
    return path


def generate_borderless_table_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        return None

    path = DIR / "borderless-table.pdf"
    doc = SimpleDocTemplate(str(path), pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph("PACKING LIST — NO BORDERS", styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Supplier: Dongguan Precision Parts Ltd.", styles["Normal"]))
    elements.append(Paragraph("Buyer: HK Trading Co.", styles["Normal"]))
    elements.append(Spacer(1, 8))

    data = [
        ["Item", "Description", "Qty", "Unit", "Weight/kg"],
        ["A001", "Aluminium casings for tablets", "2000", "pcs", "450.0"],
        ["A002", "Stainless steel screws M3x12", "10000", "pcs", "22.5"],
        ["A003", "Rubber gaskets 50x30mm", "5000", "pcs", "18.0"],
        ["B001", "LED backlight strips 55 inch", "800", "units", "160.0"],
        ["B002", "Flex PCB connectors", "3000", "pcs", "12.0"],
    ]
    table = Table(data, colWidths=[44, 190, 50, 44, 60])
    style = TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ])
    table.setStyle(style)
    elements.append(table)
    elements.append(Spacer(1, 12))
    elements.append(Paragraph("Total items: 5 | Gross weight: 662.5 kg", styles["Normal"]))
    doc.build(elements)
    print(f"Generated: {path}")
    return path


def generate_packing_list_csv():
    path = DIR / "packing-list.csv"
    rows = [
        ["item_no", "description", "hs_code", "quantity", "unit", "unit_price_usd", "total_value_usd", "gross_weight_kg", "country_of_origin"],
        ["001", "Cotton knitted pullovers", "6110.20.00", "2000", "pcs", "15.00", "30000.00", "400.0", "China"],
        ["002", "Portable digital computers <=10kg", "8471.30.00", "150", "units", "920.00", "138000.00", "225.0", "China"],
        ["003", "Electronic integrated circuits", "8542.31.00", "5000", "units", "8.50", "42500.00", "25.0", "Taiwan"],
        ["004", "Medicaments in measured doses", "3004.90.00", "1000", "bottles", "45.00", "45000.00", "180.0", "India"],
        ["005", "Electrocardiographs", "9018.11.00", "50", "units", "3200.00", "160000.00", "350.0", "Germany"],
    ]
    with open(path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    print(f"Generated: {path}")
    return path


def generate_wechat_scan():
    path = DIR / "wechat-scan.txt"
    content = """WECHAT SCREENSHOT — OCR EXTRACTED TEXT
───────────────────────────────────────

[WeChat Chat — 14 Jul 2026 14:32]

小李: 张总，今天发货单号是多少？帮我查一下。
张总: 等下，我拍给你。

— Image Description: Handwritten cargo manifest on notebook paper —

Extracted text (OCR confidence: 78%):
"发件人: 东莞华强电子
 收货人: 香港捷运物流
 货柜: OOLU8125479
 货物:
 1. 电子元器件 500箱 HS Code: 85423100 重量: 250kg
 2. LED 显示屏 200箱 重量: 180kg
 总价值: 约 USD 85,000
 目的地: 香港 → 转口 越南"

[Image Description ends]
"""
    path.write_text(content)
    print(f"Generated: {path}")
    return path


def generate_invoice_xlsx():
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl not available, skipping xlsx")
        return None

    path = DIR / "invoice.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    ws["A1"] = "INVOICE"
    ws["A2"] = "INV-2026-0715-0099"
    ws["A3"] = "Date: 15-JUL-2026"
    ws["A5"] = "Consignor: Guangzhou Tech Parts Co."
    ws["A6"] = "Consignee: Macau Logistics Ltd."
    ws["A8"] = "Port of Loading: Nansha, Guangzhou"
    ws["B8"] = "Port of Discharge: Macau"
    ws["A9"] = "Container: MSCU9032741"
    ws["A10"] = "Incoterms: CIF Macau"
    ws["A12"] = "Item"
    ws["B12"] = "Description"
    ws["C12"] = "HS Code"
    ws["D12"] = "Quantity"
    ws["E12"] = "Unit"
    ws["F12"] = "Unit Price USD"
    ws["G12"] = "Total USD"
    ws["H12"] = "Weight kg"

    items = [
        [1, "Lithium-ion battery packs 48V/20Ah", "8507.60.00", 500, "units", 120.00, 60000.00, 750.0],
        [2, "Electric motors for drones, 1000W", "8501.40.00", 200, "units", 450.00, 90000.00, 320.0],
        [3, "Carbon fiber propellers 15 inch", "8803.90.00", 2000, "pcs", 8.50, 17000.00, 40.0],
    ]
    for i, row in enumerate(items, 13):
        for j, val in enumerate(row):
            ws.cell(row=i, column=j + 1, value=val)

    ws["A17"] = "TOTAL DECLARED VALUE"
    ws["G17"] = 167000.00
    ws["A18"] = "TOTAL WEIGHT"
    ws["H18"] = 1110.0

    wb.save(str(path))
    print(f"Generated: {path}")
    return path


def generate_plain_text():
    path = DIR / "plain-invoice.txt"
    content = """INVOICE NO: INV-TEST-001
DATE: 2026-07-17

From: Test Supplier Ltd., Unit 4, 123 Queen's Road East, Wan Chai, Hong Kong
To: Test Buyer Inc., 456 Nathan Road, Tsim Sha Tsui, Kowloon, Hong Kong

Item 1: Test integrated circuit boards
  HS Code: 8542.31.00
  Quantity: 1000 units
  Unit Price: HKD 15.00
  Total: HKD 15,000.00
  Weight: 25 KG
  Origin: CN

Item 2: Test LED displays
  HS Code: 8541.41.00
  Quantity: 500 units
  Unit Price: HKD 220.00
  Total: HKD 110,000.00
  Weight: 180 KG
  Origin: CN

Container: TEST1234567
Port of Loading: Hong Kong
Port of Discharge: Shenzhen
Incoterms: FOB Hong Kong

Total Value: HKD 125,000.00
Total Weight: 205 KG
"""
    path.write_text(content)
    print(f"Generated: {path}")
    return path


if __name__ == "__main__":
    generate_invoice_pdf()
    generate_borderless_table_pdf()
    generate_packing_list_csv()
    generate_wechat_scan()
    generate_invoice_xlsx()
    generate_plain_text()
    print(f"\nAll files generated in {DIR}")
